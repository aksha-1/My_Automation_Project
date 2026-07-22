import openpyxl
from openpyxl import load_workbook
import os
import json
class DDT:
    @staticmethod
    def get_login_data():
        workbook = load_workbook(os.path.join(os.getcwd(), "TestData", "DDT.xlsx"))
        sheet = workbook["Sheet1"]
        data = []
        for row in sheet.iter_rows(min_row=2,values_only=True):
            data.append(row)
        return data
    def read_json_data():
        path=os.path.join(os.getcwd(),"TestData","ddt_json.json")
        with open (path,"r") as file:
            json_file=json.load(file)
            file1=[]
            for dict in json_file:
                file1.append((dict['username'],dict['password'],dict['expected']))
            return file1

    

